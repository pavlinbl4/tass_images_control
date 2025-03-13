from datetime import datetime
from pathlib import Path

from loguru import logger
from openpyxl import Workbook, load_workbook


class XlsxTools:

    def __init__(self, file_name):
        self.base_dir = Path().home() / 'Library/Mobile Documents/com~apple~CloudDocs/Documents/TASS/Images_on_site'
        self.file_path = self.base_dir / file_name

    def initialize(self):
        """Создает файл и лист, если их нет"""
        try:
            self.file_path.parent.mkdir(parents=True, exist_ok=True)  # Создание папки
            if not self.file_path.exists():
                self.create_file()
            else:
                logger.debug("File already exists added new sheet")
                workbook = load_workbook(filename=self.file_path, read_only=False)
                self.create_sheet(workbook)
        except Exception as e:
            logger.error(f"Ошибка при инициализации: {e}")
        return self.file_path

    def create_file(self):
        """Создает новый xlsx-файл с листом, названным по дате"""
        try:
            workbook = Workbook()
            self.create_sheet(workbook)
            logger.info(f"Создан новый файл: {self.file_path}")
        except Exception as e:
            logger.error(f"Ошибка при создании файла: {e}")

    def create_sheet(self, workbook):
        title = datetime.now().strftime("%Y%m%d-%H%M")
        sheet = workbook.create_sheet(title)
        headers = ['images_online', 'image_id', 'image_date', 'image_caption', 'image_link']
        sheet.append(headers)
        # Установка ширины колонок
        column_widths = [5, 10, 10, 30, 50]  # Примерные ширины колонок
        for col_num, width in enumerate(column_widths, start=1):
            sheet.column_dimensions[sheet.cell(row=1, column=col_num).column_letter].width = width
        workbook.active = workbook[title]
        workbook.save(self.file_path)

    def append_data(self, file_path, data):
        """Добавляет строку данных в существующий лист, созданный при первом обращении"""
        try:
            if not file_path.exists():
                logger.warning(f"Файл {file_path} не найден, создаю новый.")
                self.create_file()

            workbook = load_workbook(file_path)

            sheet = workbook.active
            sheet.append(data)

            workbook.save(file_path)
            logger.info(f"Добавлены данные в {file_path}")
        except Exception as e:
            logger.error(f"Ошибка при добавлении данных: {e}")


if __name__ == '__main__':
    xlsx_file = XlsxTools('test.xlsx').initialize()

    XlsxTools('test.xlsx').append_data(xlsx_file, ["Yes", 123, "2025-03-11", "Test Caption", "http://example.com"])
